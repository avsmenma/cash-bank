import openpyxl
wb = openpyxl.load_workbook('TemplateBank.xlsx')
print('Sheet names:', wb.sheetnames)
for i, ws in enumerate(wb.worksheets):
    print(f'\n=== Sheet {i}: {ws.title} ===')
    for j, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
        print(f'Row {j}: {list(row)}')
